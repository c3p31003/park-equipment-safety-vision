from flask import Flask, render_template,send_file, request, jsonify, redirect, url_for, session
from flask_migrate import Migrate
from werkzeug.security import check_password_hash
from models import (
    db, User, Park, Equipment, Inspection, 
    InspectionDetail, InspectionPhoto, DailyReportPhoto,
    InspectionPartEnum, TypeOfAbnormalityEnum, GradeEnum
)
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
# from flask_cors import CORS
from config import DATABASE_URL
import os
import sys
import base64
from datetime import datetime
from keras.models import load_model
from keras.preprocessing import image
from PIL import Image
import io
import json
import numpy as np

from degradation_main import run_inference




EMU = 9525
ICON_PX = 16

BASE_DIR = os.path.dirname(__file__)
TEMPLATE_PATH = os.path.join(BASE_DIR, "template.xlsx")
ICON_DIR = os.path.join(BASE_DIR, "icons")


app = Flask(__name__)
# GitHub Pages からのアクセスを許可
# CORS(app, resources={r"/api/*": {"origins": "*"}})

app.config['SECRET_KEY'] = 'your-fixed-secret-key-change-this-in-production'
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
migrate = Migrate(app, db)






# ============================================================
# モデル読み込み（改善版：4つのパーツ対応）
# ============================================================

MODELS_CONFIG = {
    'chain': {
        'path': './chain_best.keras',
        'size': 224,
        'classes': ['normal', 'rust_B', 'rust_C']
    },
    'joint': {
        'path': './joint_best.keras',
        'size': 224,
        'classes': ['normal', 'rust_B', 'rust_C']
    },
    'pole': {
        'path': './pole_best.keras',
        'size': 224,
        'classes': ['normal', 'rust_B', 'rust_C']
    },
    'seat': {
        'path': './seat_best.keras',
        'size': 224,
        'classes': ['normal', 'rust_B', 'rust_C', 'crack_B', 'crack_C']
    }
}

inference_models = {}


#ログイン機能


@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        try:
            employee_id = request.form.get('employee_id')
            user_password = request.form.get('password')

            sys.stderr.write(f"\nフォーム入力 - employee_id: {employee_id}, password: {user_password}\n")
            sys.stderr.flush()

            user = None
            try:
                sys.stderr.write(f"DB検索開始: employee_id={int(employee_id)}\n")
                sys.stderr.flush()
                user = User.query.filter_by(employee_id=int(employee_id)).first()
                sys.stderr.write(f"DB検索結果: user={user}\n")
                if user:
                    sys.stderr.write(f"  DB employee_id: {user.employee_id}\n")
                    sys.stderr.write(f"  DB password: {user.password}\n")
                sys.stderr.flush()
            except Exception as e:
                sys.stderr.write(f"クエリエラー: {e}\n")
                sys.stderr.flush()

            sys.stderr.write(f"認証判定:\n")
            sys.stderr.write(f"  user is not None: {user is not None}\n")
            if user:
                sys.stderr.write(f"  パスワード一致: {user_password} == {user.password} → {user_password == user.password}\n")
            sys.stderr.flush()

            if user and user_password == user.password:
                session['user_id'] = user.employee_id
                session['user_password'] = user.password
                session['user_name'] = user.name
                sys.stderr.write(f"認証成功！ /home にリダイレクト\n")
                sys.stderr.flush()
                return redirect(url_for('home'))
            else:
                sys.stderr.write(f"認証失敗\n")
                sys.stderr.flush()
                return render_template('Login.html', error="ユーザー名またはパスワードが正しくありません")
        except Exception as e:
            sys.stderr.write(f"エラー発生: {e}\n")
            sys.stderr.flush()
            return render_template('Login.html', error="エラーが発生しました。再度お試しください。")

    return render_template('Login.html')


@app.route('/home', methods=['GET'])
def home():
    user_name = session.get('user_name')
    if user_name:
        return render_template('index.html', employee_id=session.get('user_id'), user_name=user_name)
    else:
        return redirect(url_for('login'))





# ---------------------------帳票機能---------------------------
# ---------------------------
# TEXT 用の関数
# ---------------------------
def insert_text(ws, cell, value):
    ws[cell] = value
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top")


# ---------------------------
# ICON 挿入関数（正しい版）
# ---------------------------
def insert_icon(ws, cell, icon_file, dx=0, dy=0):
    img_path = os.path.join(ICON_DIR, icon_file)
    if not os.path.exists(img_path):
        return

    img = Image(img_path)
    img.width = ICON_PX
    img.height = ICON_PX

    # セル位置
    col_letter = ''.join(filter(str.isalpha, cell))
    row_number = int(''.join(filter(str.isdigit, cell)))
    col_idx = column_index_from_string(col_letter) - 1

    marker = AnchorMarker(
        col=col_idx,
        colOff=dx * EMU,
        row=row_number - 1,
        rowOff=dy * EMU
    )

    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(EMU * img.width, EMU * img.height)
    )

    ws.add_image(img)


# ---------------------------
#   Excel 生成 API
# ---------------------------
@app.route("/api/generate_excel", methods=["POST"])
def generate_excel():
    data = request.get_json(silent=True)
    if data is None:
        return jsonify({"error": "JSONが正しく送信されていません"}), 400

    if not os.path.exists(TEMPLATE_PATH):
        return jsonify({"error": "テンプレートファイルが見つかりません"}), 500

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    for item in data.get("items", []):
        cell = item.get("cell")
        if not cell:
            continue

        item_type = item.get("type")
        dx = item.get("dx", 0)
        dy = item.get("dy", 0)

        if item_type == "icon" and item.get("icon"):
            insert_icon(ws, cell, item["icon"], dx=dx, dy=dy)

        elif item_type in ("text", "number"):
            insert_text(ws, cell, str(item.get("value", "")))

        elif item["type"] == "text":
            cell = ws[item["cell"]]
            cell.value = item["text"]

        elif item_type == "checkbox":
            if item.get("value"):
                insert_icon(ws, cell, item.get("icon", "check.png"), dx=dx, dy=dy)

    stream = io.BytesIO()
    output_path = "backend/output.xlsx"
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="点検チェックシート.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )





def load_all_models():
    """4つのモデルを読み込む"""
    global inference_models
    
    for part_name, config in MODELS_CONFIG.items():
        try:
            model = load_model(config['path'])
            inference_models[part_name] = model
            print(f"✓ {part_name} モデル読み込み成功: {config['path']}")
        except Exception as e:
            print(f"⚠ {part_name} モデル読み込み失敗: {e}")
            inference_models[part_name] = None

# 起動時にモデルを読み込む
load_all_models()

# ============================================================
# 推論関数（改善版）
# ============================================================

def predict_equipment_part(image_binary, part_name):
    """
    画像から指定されたパーツを推論
    
    Args:
        image_binary: バイナリ画像データ
        part_name: 'chain', 'joint', 'pole', 'seat'
    
    Returns:
        (predicted_class, confidence, all_confidences)
        例：('rust_B', 0.85, {'normal': 0.05, 'rust_B': 0.85, 'rust_C': 0.10})
    """
    
    if part_name not in inference_models or inference_models[part_name] is None:
        return None, None, None
    
    try:
        # 画像をPIL Imageに変換
        img = Image.open(io.BytesIO(image_binary))
        
        # RGB に変換
        if img.mode != 'RGB':
            img = img.convert('RGB')
        
        # モデルに合わせてリサイズ
        img_size = MODELS_CONFIG[part_name]['size']
        img = img.resize((img_size, img_size))
        
        # 正規化
        img_array = np.array(img, dtype='float32') / 255.0
        img_batch = np.expand_dims(img_array, axis=0)
        
        # 予測実行
        model = inference_models[part_name]
        predictions = model.predict(img_batch, verbose=0)
        
        # 結果を取得
        confidence = float(np.max(predictions[0]))
        predicted_class_index = int(np.argmax(predictions[0]))
        predicted_class = MODELS_CONFIG[part_name]['classes'][predicted_class_index]
        
        # すべてのクラスの信頼度
        all_confidences = {
            MODELS_CONFIG[part_name]['classes'][i]: float(predictions[0][i])
            for i in range(len(MODELS_CONFIG[part_name]['classes']))
        }
        
        sys.stderr.write(f"✓ {part_name} 予測: {predicted_class} ({confidence:.4f})\n")
        sys.stderr.flush()
        
        return predicted_class, confidence, all_confidences
    
    except Exception as e:
        sys.stderr.write(f"✗ {part_name} 予測エラー: {e}\n")
        sys.stderr.flush()
        return None, None, None

def class_to_condition(predicted_class):
    """
    予測クラスを TypeOfAbnormalityEnum に変換
    
    例：
        'normal' → TypeOfAbnormalityEnum.NORMAL
        'rust_B' → TypeOfAbnormalityEnum.RUST
        'crack_B' → TypeOfAbnormalityEnum.CRACK
    """
    if 'rust' in predicted_class.lower():
        return TypeOfAbnormalityEnum.RUST
    elif 'crack' in predicted_class.lower():
        return TypeOfAbnormalityEnum.CRACK
    else:
        return TypeOfAbnormalityEnum.NORMAL

def class_to_grade(predicted_class):
    """
    予測クラスを Grade に変換
    
    例：
        'normal' → GradeEnum.A
        'rust_B' → GradeEnum.B
        'rust_C' → GradeEnum.C
    """
    if 'normal' in predicted_class.lower():
        return GradeEnum.A
    elif 'B' in predicted_class:
        return GradeEnum.B
    elif 'C' in predicted_class:
        return GradeEnum.C
    else:
        return GradeEnum.A

def part_name_to_enum(part_name):
    """文字列をInspectionPartEnumに変換"""
    part_map = {
        'chain': InspectionPartEnum.CHAIN,
        'joint': InspectionPartEnum.JOINT,
        'pole': InspectionPartEnum.POLE,
        'seat': InspectionPartEnum.SEAT
    }
    return part_map.get(part_name, InspectionPartEnum.CHAIN)



@app.route('/CheckSheet')
def CheckSheet():
    return render_template('CheckSheet.html')

@app.route('/daily_report')
def daily_report():
    return render_template('daily_report.html')

@app.route('/inspection_results')
def inspection_results():
    return render_template('inspection_results.html')

@app.route('/AllDocuments')
def AllDocuments():
    return render_template('AllDocuments.html')

@app.route('/PhotoViewing')
def PhotoViewing():
    return render_template('PhotoViewing.html')

@app.route('/TakePhoto')
def TakePhoto():
    return render_template('TakePhoto.html')

@app.route('/results_report')
def results_report():
    return render_template('results_report.html')

@app.route('/Deterioration')
def Deterioration():
    return render_template('Deterioration.html')


# ============================================================
# API エンドポイント（改善版：4パーツ対応）
# ============================================================

@app.route('/api/inspection/<int:inspection_id>/upload_photo', methods=['POST'])
def upload_photo(inspection_id):
    """
    写真アップロード + AI判定結果保存（4パーツ対応版）
    
    Request JSON:
    {
        "photo_data": "<base64_image>",
        "filename": "inspection_001.jpg",
        "parts": {
            "chain": {"image_data": "<base64>"},
            "joint": {"image_data": "<base64>"},
            "pole": {"image_data": "<base64>"},
            "seat": {"image_data": "<base64>"}
        }
    }
    """
    
    try:
        data = request.json
        
        # 1. 点検レコードを取得
        inspection = Inspection.query.get_or_404(inspection_id)
        
        # 2. 各パーツについて推論と結果保存
        part_results = {}
        worst_grade = GradeEnum.A
        
        if 'parts' in data:
            for part_name, part_data in data['parts'].items():
                if not part_data or 'image_data' not in part_data:
                    continue
                
                try:
                    # Base64デコード
                    image_base64 = part_data['image_data']
                    if ',' in image_base64:
                        image_base64 = image_base64.split(',')[1]
                    
                    image_binary = base64.b64decode(image_base64)
                    
                    # 推論実行
                    predicted_class, confidence, all_confidences = predict_equipment_part(
                        image_binary, 
                        part_name
                    )
                    
                    if predicted_class is None:
                        part_results[part_name] = {
                            'error': 'Model not loaded or prediction failed'
                        }
                        continue
                    
                    # Condition と Grade に変換
                    condition = class_to_condition(predicted_class)
                    grade = class_to_grade(predicted_class)
                    part_enum = part_name_to_enum(part_name)
                    
                    # 3. InspectionDetail レコードを取得または作成
                    detail = InspectionDetail.query.filter_by(
                        inspection_id=inspection_id,
                        part=part_enum
                    ).first()
                    
                    if detail:
                        detail.condition = condition
                        detail.grade = grade
                        detail.confidence = confidence
                        detail.is_ai_predicted = True
                        detail.updated_at = datetime.utcnow()
                    else:
                        detail = InspectionDetail(
                            inspection_id=inspection_id,
                            part=part_enum,
                            condition=condition,
                            grade=grade,
                            confidence=confidence,
                            is_ai_predicted=True,
                            ai_raw_result=json.dumps({
                                'part': part_name,
                                'predicted_class': predicted_class,
                                'confidence': confidence,
                                'all_confidences': all_confidences,
                                'timestamp': datetime.utcnow().isoformat()
                            })
                        )
                        db.session.add(detail)
                    
                    db.session.flush()
                    
                    # 4. Photo レコード作成
                    photo = InspectionPhoto(
                        inspection_id=inspection_id,
                        detail_id=detail.detail_id,
                        photo_data=image_binary,
                        file_size=len(image_binary),
                        uploaded_by=session.get('user_id')
                    )
                    db.session.add(photo)
                    db.session.flush()
                    
                    # 5. 結果を保存
                    part_results[part_name] = {
                        'success': True,
                        'detail_id': detail.detail_id,
                        'photo_id': photo.photo_id,
                        'predicted_class': predicted_class,
                        'confidence': float(confidence),
                        'grade': grade.value,
                        'condition': condition.value
                    }
                    
                    # 最悪グレードを更新
                    grade_order = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
                    if grade_order.get(grade.value, 0) > grade_order.get(worst_grade.value, 0):
                        worst_grade = grade
                    
                except Exception as part_error:
                    sys.stderr.write(f"❌ {part_name} 処理エラー: {str(part_error)}\n")
                    sys.stderr.flush()
                    part_results[part_name] = {'error': str(part_error)}
        
        # 6. Inspection テーブルを更新
        inspection.photography_at = datetime.utcnow()
        inspection.photographer_id = session.get('user_id')
        inspection.overall_grade = worst_grade
        
        # 7. コミット
        db.session.commit()
        
        sys.stderr.write(f"✓ 点検ID {inspection_id} の処理完了\n")
        sys.stderr.flush()
        
        # 8. レスポンス
        return jsonify({
            'success': True,
            'inspection_id': inspection_id,
            'overall_grade': worst_grade.value,
            'parts': part_results,
            'timestamp': datetime.utcnow().isoformat()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        sys.stderr.write(f"❌ エラー: {str(e)}\n")
        sys.stderr.flush()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/inspection/<int:inspection_id>/results', methods=['GET'])
def get_inspection_results(inspection_id):
    """点検結果を取得（既存コード）"""
    try:
        from models import Inspection, InspectionDetail
        
        inspection = Inspection.query.get_or_404(inspection_id)
        details = InspectionDetail.query.filter_by(inspection_id=inspection_id).all()
        
        results = {
            'inspection_id': inspection_id,
            'overall_grade': inspection.overall_grade.value if inspection.overall_grade else None,
            'parts': []
        }
        
        for detail in details:
            results['parts'].append({
                'part': detail.part.value,
                'condition': detail.condition.value if detail.condition else None,
                'grade': detail.grade.value if detail.grade else None,
                'confidence': detail.confidence,
                'is_ai_predicted': detail.is_ai_predicted
            })
        
        return jsonify(results)
        
    except Exception as e:
        sys.stderr.write(f"❌ エラー: {str(e)}\n")
        sys.stderr.flush()
        return jsonify({'error': str(e)}), 500


# ============================================================
# ヘルスチェック（新規追加：モデル状態確認用）
# ============================================================

@app.route('/api/health', methods=['GET'])
def health():
    """推論エンジンのステータス確認"""
    
    models_status = {}
    for part_name in MODELS_CONFIG.keys():
        if inference_models.get(part_name) is not None:
            models_status[part_name] = 'loaded'
        else:
            models_status[part_name] = 'not_loaded'
    
    all_loaded = all(v == 'loaded' for v in models_status.values())
    
    return jsonify({
        'status': 'ok' if all_loaded else 'partial',
        'models': models_status,
        'timestamp': datetime.utcnow().isoformat()
    }), 200


# ～劣化診断機能～
# HTML/JS からの写真アップロード → 劣化度を返す API
@app.route("/api/degradation", methods=["POST"])
def api_degradation():
    """
    HTML からアップロードされた写真を受け取り、
    run_inference で劣化度を計算して返す
    """
    file = request.files.get("photo")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    # 一時保存用フォルダ
    tmp_dir = "data/raw"
    os.makedirs(tmp_dir, exist_ok=True)
    tmp_path = os.path.join(tmp_dir, file.filename)
    file.save(tmp_path)

    try:
        # run_inference を呼ぶ
        degradation_ratio, _, _ = run_inference(tmp_path)

        # % に変換して返す
        return jsonify({"degradation_ratio": round(degradation_ratio * 100, 2)})

    except Exception as e:
        return jsonify({"error": str(e)}), 500





# ============================================================
# DB初期化（既存コード）
# ============================================================

tables_created = os.path.exists('db_initialized.flag')

if not tables_created:
    with app.app_context():
        db.create_all()
        print("テーブルが作成されました")
        
        # テストユーザーを作成
        try:
            test_user = User.query.filter_by(employee_id=1).first()
            if not test_user:
                test_user = User(
                    employee_id=1,
                    name="テストユーザー",
                    password="1234",
                    role="STAFF"
                )
                db.session.add(test_user)
                db.session.commit()
                print("✓ テストユーザー作成: employee_id=1, password=1234")
        except Exception as e:
            print(f"⚠ テストユーザー作成失敗: {e}")
        
        with open('db_initialized.flag', 'w') as f:
            f.write('created')
    print("データベースが初期化されました")


if __name__ == '__main__':
    app.run(debug=True)